# -*- coding:utf-8 -*-
import openpyxl
import os
import shutil
import datetime
import numpy as np
import pandas as pd

"""
create a excel file, and move to the path of the specified.
first, judge the exists of the file;
second, create a new blank file;
finally, move the file to the path.
"""


def create_workbook(workbook_name, path):
    file_full_name = path + workbook_name
    if os.path.exists(file_full_name):
        # print("The file %s exists, can't create a new one." % file_full_name)
        return False
    else:
        create_workbook_with_single_sheet(workbook_name, path, None)
        return True


def create_workbook_with_single_sheet(workbook_name, path, sheet_name):
    file_full_name = path + workbook_name
    if os.path.exists(file_full_name):
        # print("The file %s exists, can't create a new one." % file_full_name)
        return False
    else:
        workbook = openpyxl.Workbook()
        if sheet_name:
            worksheet = workbook.active
            worksheet.title = sheet_name
        workbook.save(filename=workbook_name)
        shutil.move(workbook_name, file_full_name)
        return True


"""
read single sheet's all data.
"""


def read_row_data(worksheet):
    for row in worksheet.rows:
        row_data = []
        for cell in row:
            cell_value = cell.value
            # 空数据暂不处理，由pandas统一处理
            if not cell_value:
                cell_value = np.NaN
            row_data.append(cell_value)
        yield row_data


def read_single_sheet_all_data(worksheet):
    sheet_data = []
    # print(worksheet['B4'].number_format)  # test
    for row_data in read_row_data(worksheet):
        if row_data:
            sheet_data.append(row_data)
    return sheet_data


"""
read an exist excel file.
first, check the exist of the file;
then, if it exists, load it, and read all of specific sheet data.
As you see, the sheet_name may be more than one.
So, the return value data is multiple type, like
{'sheet_name1': 'sheet_data1', 'sheet_name2': 'sheet_data2'}
etc...
"""


def read_workbook(workbook_name, path="", sheet_name=""):
    file_full_name = path + workbook_name
    if os.path.exists(file_full_name):
        workbook = openpyxl.load_workbook(filename=file_full_name, data_only=True)
        all_worksheet_name = workbook.get_sheet_names()
        # 修订，若sheet name 未指定，则为所有sheet
        if sheet_name == "":
            sheet_name = all_worksheet_name
        data = {}
        if sheet_name == "":
            worksheet = workbook.active
            sheet_data = read_single_sheet_all_data(worksheet)
            data.update({worksheet.title: sheet_data})
        else:
            for single_sheet_name in sheet_name:
                if single_sheet_name not in all_worksheet_name:
                    continue
                worksheet = workbook[single_sheet_name]
                sheet_data = read_single_sheet_all_data(worksheet)
                data.update({single_sheet_name: sheet_data})
        return data


def read_workbook_range_in_single_sheet(workbook_name, path="", single_sheet_name="Sheet1", range='A1'):
    file_full_name = path + workbook_name
    if os.path.exists(file_full_name):
        workbook = openpyxl.load_workbook(filename=file_full_name, data_only=True)
        ws = workbook[single_sheet_name]
        return ws[range]



"""
清洗读取的原始数据
"""


def data_clean_in_one_sheet(org_data, sheet_name, row_start_num=1, field_row_num=0, axis_zero_na_thresh=5,
                            axis_one_na_thresh=5, axis_zero_drop_na_flag=True, axis_one_drop_na_flag=True):
    trade_data_df = pd.DataFrame(org_data.get(sheet_name)[row_start_num:],
                                 columns=org_data.get(sheet_name)[field_row_num])
    if axis_zero_drop_na_flag:
        trade_data_df = trade_data_df.dropna(axis=0, how='any', thresh=axis_zero_na_thresh)
    if axis_one_drop_na_flag:
        trade_data_df = trade_data_df.dropna(axis=1, how='any', thresh=axis_one_na_thresh)
    trade_data_df = trade_data_df.fillna(value="")
    return trade_data_df


"""
save data to specific workbook.
As you see, the data structrue is
[{"sheet_name": "sheet_name", "sheet_data": data, "fields": fields,
  --- thinking in the last attr
  --- "row_offset": number or None(None is the maxrow),
  --- "collum_offset": number or None,
  --- "save_type": "append" or "cover" or "repeat"(default)
  }]
, so
the sheetname and other messages is contained in the data.
if the sheet name is duplicate, add '_copy' string and now time string
after the original sheet name.
if there's no this sheet, then create one.
"""


def save_data_to_workbook(info, workbook_name, path="", over_write=True):
    file_full_name = path + workbook_name
    if not os.path.exists(file_full_name):
        create_workbook(workbook_name, path)
    workbook = openpyxl.load_workbook(filename=file_full_name, guess_types = True)
    all_sheet_name = workbook.get_sheet_names()
    for sheet_info in info:
        sheet_name = sheet_info.get("sheet_name")
        if sheet_name == "":
            sheet_name = "iNeedAName"
        if sheet_name in all_sheet_name and not over_write:
            now_time_str = datetime.datetime.now().strftime("%H%M%S")
            sheet_name = sheet_name + "_copy" + now_time_str
        worksheet = workbook.create_sheet(sheet_name, 0)
        # worksheet.title = sheet_name
        # now copy the data
        fields = sheet_info.get("fields")
        sheet_data = sheet_info.get("sheet_data")
        # print(sheet_data)
        if fields is not None:
            worksheet.append(fields)
        for row_data in sheet_data:
            worksheet.append(row_data)

    workbook.save(filename=file_full_name)


"""
remove a file
"""


def remove_file(filename, path):
    if os.path.exists(path + filename):
        os.remove(path + filename)
        return True
    else:
        return False


def read_number_format_from_cell(filename, path, sheet_name, cell_str):
    file_full_name = path + filename
    if os.path.exists(file_full_name):
        workbook = openpyxl.load_workbook(filename = file_full_name)
        worksheet = workbook[sheet_name]
        cell = worksheet[cell_str]
        return cell.number_format


"""
apply the number_format to the whole column:
常规: General, the default value
数值: 0.00_);[Red]\(0.00\)
千分位数值：#,##0.00_)
百分比: 0.00%
短日期: mm-dd-yy
info's data structrue is :
{"columnA": "formatA"...}
"""


def apply_number_format(filename, path, sheet_name, info):
    file_full_name = path + filename
    if os.path.exists(file_full_name):
        wb = openpyxl.load_workbook(file_full_name)
        ws = wb[sheet_name]

        for info_key in info.keys():
            # print(info_key)
            # print(ws['C'])
            # print(ws[1])
            for cell in ws[info_key]:
                cell.number_format = info[info_key]

        wb.save(filename=file_full_name)


def apply_column_width(filename, path, sheet_name, info):
    file_full_name = path + filename
    if os.path.exists(file_full_name) and info:
        wb = openpyxl.load_workbook(file_full_name)
        ws = wb[sheet_name]

        for info_key in info.keys():
            ws.column_dimensions[info_key].width = info[info_key]

        wb.save(filename=file_full_name)


# test fragment
if __name__ == "__main__":
    test = read_workbook_range_in_single_sheet("d:/投资收益月度核对/期初资本公积.xlsx").value
    print(test / 100)
    # create_workbook("test.xlsx", "d:/")
    # original_data = read_workbook("格式测试.xlsx", "d:/", ['Sheet1'])
    # info = []
    # for key in original_data.keys():
    #     sheet_info = {}
    #     sheet_info.update({"sheet_name": key})
    #     sheet_info.update({"fields": ["1", "2"]})
    #     sheet_info.update({"sheet_data": original_data[key]})
    #     info.append(sheet_info)
    # save_data_to_workbook(info, "test2.xlsx", "d:/")
    # apply_number_format("test2.xlsx", "d:/", "Sheet",
    #                      {"B": "0.00_);[Red]\(0.00\)",
    #                       "C": "0.00%",
    #                       "D": "mm-dd-yy"})
